"""
exports/excel_exporter.py — Generación del Excel de conciliación CREDIEXPRESS
Produce un archivo con 10 hojas: comparación completa, coincidencias, agrupados,
rechazos, solo banco, solo auxiliar, resumen ejecutivo y acciones requeridas.
"""
import io
import logging

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

# ── Paleta de colores por estado ──────────────────────────────────────────────
FILL_VERDE    = PatternFill("solid", fgColor="C8F7C5")
FILL_AMARILLO = PatternFill("solid", fgColor="FFF3CD")
FILL_ROJO     = PatternFill("solid", fgColor="F7C5C5")
FILL_AZUL     = PatternFill("solid", fgColor="D0E8FF")
FILL_NARANJA  = PatternFill("solid", fgColor="FFE0B2")
FILL_CELESTE  = PatternFill("solid", fgColor="B3E5FC")
FILL_HEADER   = PatternFill("solid", fgColor="1565C0")
FONT_HEADER   = Font(bold=True, color="FFFFFF", size=10)

# ── Helpers de estilo ─────────────────────────────────────────────────────────
def _estilizar_hoja(ws):
    for cell in ws[1]:
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col in ws.columns:
        mx = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 3, 55)

def _colorear_por_estado(ws, col_estado_idx):
    for row in ws.iter_rows(min_row=2):
        val = str(row[col_estado_idx - 1].value or "")
        fill = (FILL_VERDE    if "COINCIDE EXACTO"  in val else
                FILL_AMARILLO if "COINCIDE APROX"   in val else
                FILL_CELESTE  if "AGRUPADO"         in val else
                FILL_NARANJA  if "RECHAZO"          in val else
                FILL_ROJO     if "SOLO EN BANCO"    in val else
                FILL_AZUL     if "SOLO EN AUXILIAR" in val else None)
        if fill:
            for cell in row:
                cell.fill = fill

# ── Función principal ─────────────────────────────────────────────────────────
def generar_excel(
    df_comp, df_banco, df_aux, df_solo_aux,
    banco_nombre, aux_nombre,
    sa, sac, tab_s, tca_s,
    si_a, sf_a, td_a, tc_a,
    n_tot, n_exac, n_apr, n_agr, n_rec, n_sbco, n_saux, pct_conc,
    nombre_salida="CREDIEXPRESS_Conciliacion.xlsx",
):
    """
    Genera el Excel de conciliación y devuelve bytes (io.BytesIO).

    Parámetros
    ----------
    df_comp        : DataFrame resultado de comparar_documentos()
    df_banco       : DataFrame extracto bancario original
    df_aux         : DataFrame auxiliar contable original
    df_solo_aux    : DataFrame con asientos sin movimiento bancario
    banco_nombre   : str — nombre del archivo banco
    aux_nombre     : str — nombre del archivo auxiliar
    sa, sac        : float — saldo inicial/final banco
    tab_s, tca_s   : float — total abonos/cargos banco
    si_a, sf_a     : float — saldo inicial/final auxiliar
    td_a, tc_a     : float — total débitos/créditos auxiliar
    n_tot, n_exac, n_apr, n_agr, n_rec, n_sbco, n_saux : int — contadores
    pct_conc       : float — tasa de conciliación %
    nombre_salida  : str  — nombre sugerido del archivo
    """
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # ── Hoja 1: Comparación completa ─────────────────────────────────────
        if not df_aux.empty:
            h1 = df_comp[["N","FECHA_BANCO","TIPO_MOV","DESCRIPCION","VALOR_BANCO",
                           "DOC_AUXILIAR","FECHA_AUXILIAR","CONCEPTO_AUX","MONTO_AUXILIAR",
                           "DIFERENCIA","ESTADO"]].copy()
            h1.columns = ["N","Fecha_Banco","Tipo","Descripcion_Banco","Valor_Banco",
                          "Doc_Auxiliar","Fecha_Auxiliar","Concepto_Auxiliar","Monto_Auxiliar",
                          "Diferencia","Estado"]
        else:
            h1 = pd.DataFrame({"Info": ["Sin comparacion"]})
        h1.to_excel(writer, sheet_name="1_Comparacion_Completa", index=False)

        # ── Hojas 2-6: subconjuntos por estado ───────────────────────────────
        for estado, nombre in [
            ("COINCIDE EXACTO", "2_Coincidencias_Exactas"),
            ("COINCIDE APROX.", "3_Coincidencias_Aprox"),
            ("AGRUPADO",        "4_Agrupados_N1"),
            ("RECHAZO",         "5_Rechazos_Confirmar"),
            ("SOLO EN BANCO",   "6_Solo_Banco_Sin_Auxiliar"),
        ]:
            sub = (df_comp[df_comp["ESTADO"].str.contains(estado, na=False)].copy()
                   if not df_aux.empty else pd.DataFrame())
            if sub.empty:
                sub = pd.DataFrame({"Info": ["Sin registros"]})
            sub.to_excel(writer, sheet_name=nombre, index=False)

        # ── Hoja 7: Solo auxiliar ─────────────────────────────────────────────
        if not df_solo_aux.empty:
            df_solo_aux.to_excel(writer, sheet_name="7_Solo_Auxiliar_Sin_Banco", index=False)
        else:
            pd.DataFrame({"Info": ["Todos los asientos tienen movimiento bancario"]}).to_excel(
                writer, sheet_name="7_Solo_Auxiliar_Sin_Banco", index=False)

        # ── Datos completos ───────────────────────────────────────────────────
        df_banco.to_excel(writer, sheet_name="7_Extracto_Banco_Completo",  index=True)
        df_aux.to_excel(  writer, sheet_name="8_Auxiliar_Contable_Completo", index=True)

        # ── Hoja 9: Resumen ejecutivo ─────────────────────────────────────────
        resumen_data = {
            "Concepto": [
                "Archivo banco", "Archivo auxiliar",
                "Saldo inicial banco", "Saldo final banco",
                "Total abonos banco", "Total cargos banco",
                "Saldo inicial auxiliar", "Saldo final auxiliar",
                "Total debitos auxiliar", "Total creditos auxiliar",
                "Diferencia saldos finales",
                "Movimientos analizados", "Coincidencias exactas",
                "Coincidencias aprox.", "Agrupados N:1",
                "Rechazos confirmar", "Solo en banco", "Solo en auxiliar",
                "Tasa de conciliacion %",
            ],
            "Valor": [
                banco_nombre, aux_nombre,
                sa, sac, tab_s, tca_s,
                si_a, sf_a, td_a, tc_a,
                sac - sf_a,
                n_tot, n_exac, n_apr, int(n_agr), int(n_rec), n_sbco, n_saux,
                round(pct_conc, 1),
            ]
        }
        pd.DataFrame(resumen_data).to_excel(
            writer, sheet_name="9_Resumen_Conciliacion", index=False)

        # ── Hoja 10: Acciones requeridas ──────────────────────────────────────
        if not df_aux.empty:
            _sb_acciones = df_comp[
                df_comp['ESTADO'].str.contains('SOLO EN BANCO', na=False)].copy()
            if not _sb_acciones.empty:
                _cols_acc = [c for c in [
                    'FECHA_BANCO','TIPO_MOV','DESCRIPCION','VALOR_BANCO',
                    'CANDIDATO_DOC','CANDIDATO_CONCEPTO','CANDIDATO_MONTO',
                    'CANDIDATO_DIFF_PCT','RAZON_NO_MATCH'
                ] if c in _sb_acciones.columns]
                _sb_acc_out = _sb_acciones[_cols_acc].copy()
                _sb_acc_out.rename(columns={
                    'FECHA_BANCO':        'Fecha Banco',
                    'TIPO_MOV':           'Tipo',
                    'DESCRIPCION':        'Descripcion Banco',
                    'VALOR_BANCO':        'Valor Banco',
                    'CANDIDATO_DOC':      'Candidato Auxiliar',
                    'CANDIDATO_CONCEPTO': 'Concepto Candidato',
                    'CANDIDATO_MONTO':    'Monto Candidato',
                    'CANDIDATO_DIFF_PCT': 'Diferencia %',
                    'RAZON_NO_MATCH':     'Razon No Coincidio',
                }, inplace=True)

                def _accion_req(r):
                    diff = r.get('Diferencia %', None)
                    try:
                        diff = float(diff)
                    except Exception:
                        diff = 99
                    if diff <= 0.5:
                        return 'Revisar numero de documento o concepto en SIIGO'
                    elif diff <= 5:
                        return 'Verificar NC en SIIGO — posible diferencia de redondeo'
                    else:
                        return 'Crear Nota Contable en SIIGO con valor exacto del extracto'

                _sb_acc_out['ACCION REQUERIDA'] = _sb_acc_out.apply(_accion_req, axis=1)
                _sb_acc_out.to_excel(writer, sheet_name="10_Acciones_Requeridas", index=False)

                # Colorear hoja 10 según urgencia
                _ws10 = writer.book["10_Acciones_Requeridas"]
                from openpyxl.styles import PatternFill as _PF
                _fill_red = _PF(fill_type='solid', fgColor='FFCDD2')
                _fill_org = _PF(fill_type='solid', fgColor='FFE0B2')
                _fill_yel = _PF(fill_type='solid', fgColor='FFF9C4')
                _acc_col_idx = {cell.value: cell.column for cell in _ws10[1]}
                _diff_col    = _acc_col_idx.get('Diferencia %', None)
                for _wr in _ws10.iter_rows(min_row=2, max_row=_ws10.max_row):
                    try:
                        _dv = float(_wr[_diff_col - 1].value) if _diff_col else 99
                    except Exception:
                        _dv = 99
                    _fill = _fill_red if _dv > 5 else (_fill_org if _dv > 0.5 else _fill_yel)
                    for _wc in _wr:
                        _wc.fill = _fill

        # ── Estilo global + colores fila ──────────────────────────────────────
        wb = writer.book
        for sname in wb.sheetnames:
            _estilizar_hoja(wb[sname])
        if "1_Comparacion_Completa" in wb.sheetnames and not df_aux.empty:
            _colorear_por_estado(wb["1_Comparacion_Completa"], 11)

    output.seek(0)
    return output, nombre_salida
