"""
CREDIEXPRESS POPAYÁN SAS — Conciliación Bancaria Interactiva Premium
Soporte multiformato (PDF, CSV, Excel, TXT) + OCR para PDF escaneados
100% fiel al notebook original en procesamiento y reglas de negocio
"""

import streamlit as st
import re, io, warnings, os, tempfile, logging
from pathlib import Path
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import pdfplumber
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# OCR (opcional, solo se usa si las librerías están instaladas)
try:
    from pdf2image import convert_from_path
    import pytesseract
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False

warnings.filterwarnings('ignore')
pd.set_option('display.float_format', lambda x: f'{x:,.2f}')
pd.set_option('display.max_colwidth', 90)
pd.set_option('display.max_rows', 800)

st.set_page_config(page_title="Conciliación CREDIEXPRESS", page_icon="🏦", layout="wide")

# ── Helpers originales ────────────────────────────────────────────────────────
def cop(v):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return '                 N/A'
    signo = '-' if v < 0 else ' '
    return f'{signo}$ {abs(v):>18,.2f}'

def pct_bar(p, width=20):
    filled = int(p / 100 * width)
    return '[' + '█' * filled + '░' * (width - filled) + ']'

# ── Función OCR ─────────────────────────────────────────────────────────────
def ocr_pdf_page(pdf_path, page_number):
    """Devuelve el texto de una página específica usando OCR."""
    if not OCR_AVAILABLE:
        return ""
    try:
        images = convert_from_path(pdf_path, first_page=page_number, last_page=page_number)
        if images:
            return pytesseract.image_to_string(images[0], lang='spa')
    except Exception as e:
        logging.warning(f"Error OCR en página {page_number}: {e}")
    return ""

# ── Diagnóstico de legibilidad (original + OCR) ─────────────────────────────
def diagnosticar_pdf(ruta, tipo):
    resultado = {
        'archivo': ruta, 'tipo': tipo,
        'paginas_total': 0, 'paginas_con_texto': 0, 'paginas_sin_texto': 0,
        'total_chars': 0, 'total_words': 0, 'lineas_doc_encontradas': 0,
        'pct_paginas_legibles': 0.0, 'pct_estimado_datos': 0.0,
        'calidad': '', 'advertencias': [], 'ocr_usado': False
    }

    pat_doc = re.compile(r'(?:CON|CE|NC)-\d+')
    pat_mov_banco = re.compile(r'^\d{1,2}/\d{2}\s+', re.MULTILINE)

    try:
        with pdfplumber.open(ruta) as pdf:
            resultado['paginas_total'] = len(pdf.pages)
            for pag in pdf.pages:
                texto = pag.extract_text() or ''
                if len(texto.strip()) > 30:
                    resultado['paginas_con_texto'] += 1
                    resultado['total_chars'] += len(texto)
                    resultado['total_words'] += len(texto.split())
                    if tipo == 'AUXILIAR':
                        resultado['lineas_doc_encontradas'] += len(pat_doc.findall(texto))
                    else:
                        resultado['lineas_doc_encontradas'] += len(pat_mov_banco.findall(texto))
                else:
                    # Intentar OCR
                    if OCR_AVAILABLE:
                        ocr_text = ocr_pdf_page(ruta, pag.page_number)
                        if len(ocr_text.strip()) > 30:
                            resultado['paginas_con_texto'] += 1
                            resultado['total_chars'] += len(ocr_text)
                            resultado['total_words'] += len(ocr_text.split())
                            resultado['advertencias'].append(
                                f'Pág. {pag.page_number}: texto extraído con OCR')
                            resultado['ocr_usado'] = True
                            if tipo == 'AUXILIAR':
                                resultado['lineas_doc_encontradas'] += len(pat_doc.findall(ocr_text))
                            else:
                                resultado['lineas_doc_encontradas'] += len(pat_mov_banco.findall(ocr_text))
                        else:
                            resultado['paginas_sin_texto'] += 1
                            resultado['advertencias'].append(
                                f'Pág. {pag.page_number}: sin texto (imagen sin OCR disponible o ilegible)')
                    else:
                        resultado['paginas_sin_texto'] += 1
                        resultado['advertencias'].append(
                            f'Pág. {pag.page_number}: sin texto (imagen escaneada y OCR no instalado)')
    except Exception as e:
        resultado['advertencias'].append(f'Error al abrir: {e}')
        return resultado

    n_tot = resultado['paginas_total']
    n_ok  = resultado['paginas_con_texto']
    resultado['pct_paginas_legibles'] = (n_ok / n_tot * 100) if n_tot > 0 else 0

    if resultado['lineas_doc_encontradas'] > 0:
        pct_datos = min(100, resultado['pct_paginas_legibles'] * 0.98 +
                        min(2, resultado['lineas_doc_encontradas'] / 10))
    else:
        pct_datos = resultado['pct_paginas_legibles'] * 0.5

    resultado['pct_estimado_datos'] = round(pct_datos, 1)
    if pct_datos >= 95:
        resultado['calidad'] = '🟢 EXCELENTE'
    elif pct_datos >= 80:
        resultado['calidad'] = '🟡 BUENA'
    elif pct_datos >= 50:
        resultado['calidad'] = '🟠 PARCIAL'
    else:
        resultado['calidad'] = '🔴 BAJA'
    return resultado

# ── Parseo banco original (PDF) ─────────────────────────────────────────────
def limpiar_num(t):
    t = str(t or '').strip()
    if not t:
        return None
    neg = t.startswith('-') or (t.startswith('(') and t.endswith(')'))
    t = re.sub(r'[\$\(\)\s]', '', t).replace(',', '')
    try:
        v = float(t)
        return -abs(v) if neg else v
    except:
        return None

def es_fecha_banco(t):
    return bool(re.match(r'^\d{1,2}/\d{2}$', str(t or '').strip()))

def parsear_banco_pdf(ruta, usar_ocr=False):
    registros = []
    resumen = {}
    pat_res = {
        'SALDO_ANTERIOR': r'SALDO\s+ANTERIOR\s+\$?\s*([\d,\.]+)',
        'TOTAL_ABONOS'  : r'TOTAL\s+ABONOS\s+\$?\s*([\d,\.]+)',
        'TOTAL_CARGOS'  : r'TOTAL\s+CARGOS\s+\$?\s*([\d,\.]+)',
        'SALDO_ACTUAL'  : r'SALDO\s+ACTUAL\s+\$?\s*([\d,\.]+)',
    }

    with pdfplumber.open(ruta) as pdf:
        for n_pag, pag in enumerate(pdf.pages):
            # Obtener texto, con fallback a OCR si está vacío y usar_ocr=True
            texto = pag.extract_text() or ''
            if len(texto.strip()) <= 30 and usar_ocr and OCR_AVAILABLE:
                texto = ocr_pdf_page(ruta, pag.page_number)

            if n_pag == 0:
                for clave, pat in pat_res.items():
                    m = re.search(pat, texto, re.IGNORECASE)
                    if m and clave not in resumen:
                        v = m.group(1).replace(',', '')
                        resumen[clave] = limpiar_num(v)

            tabla = pag.extract_table({
                'vertical_strategy': 'lines',
                'horizontal_strategy': 'lines',
            })
            if not tabla:
                tabla = pag.extract_table()

            if tabla:
                for fila in tabla:
                    if not fila:
                        continue
                    celdas = [str(c or '').strip() for c in fila]
                    fecha_raw = next((c for c in celdas if es_fecha_banco(c)), None)
                    if not fecha_raw:
                        continue
                    nums = []
                    for c in reversed(celdas):
                        v = limpiar_num(c)
                        if v is not None:
                            nums.insert(0, v)
                        elif nums:
                            break
                    if len(nums) < 1:
                        continue
                    saldo = nums[-1]
                    valor = nums[-2] if len(nums) >= 2 else None
                    idx_f = celdas.index(fecha_raw)
                    n_num = len(nums)
                    desc = ' '.join(c for c in celdas[idx_f+1:len(celdas)-n_num]
                                     if c and not es_fecha_banco(c))
                    desc = re.sub(r'\s+', ' ', desc).strip()
                    registros.append({
                        'FECHA_RAW': fecha_raw,
                        'FECHA': pd.to_datetime('2025/' + fecha_raw,
                                        format='%Y/%d/%m', errors='coerce'),
                        'DESCRIPCION': desc,
                        'VALOR': valor,
                        'SALDO': saldo,
                        'TIPO': 'ABONO' if (valor or 0) >= 0 else 'CARGO',
                    })
            else:
                # Fallback línea a línea usando el texto (ya sea normal u OCR)
                for linea in texto.split('\n'):
                    partes = linea.strip().split()
                    if not partes or not es_fecha_banco(partes[0]):
                        continue
                    fecha_raw = partes[0]
                    nums = []; desc_p = []
                    for p in partes[1:]:
                        v = limpiar_num(p)
                        if v is not None:
                            nums.append(v)
                        elif not nums:
                            desc_p.append(p)
                    if not nums:
                        continue
                    saldo = nums[-1]
                    valor = nums[-2] if len(nums) >= 2 else nums[0]
                    registros.append({
                        'FECHA_RAW': fecha_raw,
                        'FECHA': pd.to_datetime('2025/' + fecha_raw,
                                        format='%Y/%d/%m', errors='coerce'),
                        'DESCRIPCION': ' '.join(desc_p),
                        'VALOR': valor,
                        'SALDO': saldo,
                        'TIPO': 'ABONO' if (valor or 0) >= 0 else 'CARGO',
                    })

    df = pd.DataFrame(registros)
    if df.empty:
        return df, resumen
    df = df[df['VALOR'].notna()].copy()
    df['VALOR'] = pd.to_numeric(df['VALOR'], errors='coerce')
    df = df.drop_duplicates(subset=['FECHA_RAW', 'DESCRIPCION', 'VALOR', 'SALDO'])
    df = df.sort_values('FECHA', na_position='last').reset_index(drop=True)
    df.index += 1
    return df, resumen

# ── Parseo auxiliar original (PDF) + OCR ─────────────────────────────────────
REGLAS_COL = [
    (re.compile(r'ABONO\s+A\s+PRESTAMO', re.I),            'DEBITO'),
    (re.compile(r'RENDIMIENTO|INTERES\s+AHORROS', re.I),     'DEBITO'),
    (re.compile(r'\bN\.D\.\b', re.I),                     'DEBITO'),
    (re.compile(r'\bPRESTAMO\b(?!.*ABONO)', re.I),          'CREDITO'),
    (re.compile(r'RETIRO\s+PARA\s+PAGO', re.I),             'CREDITO'),
    (re.compile(r'CANCELACION\s+NOMINA', re.I),              'CREDITO'),
    (re.compile(r'GASTO\s+BANCAR|\bN\.C\.\b', re.I),     'CREDITO'),
    (re.compile(r'IMPUESTO\s+MOVIMIENTO|GMF|4X1000', re.I),  'CREDITO'),
]

def determinar_columna(concepto, doc_code):
    for pat, col in REGLAS_COL:
        if pat.search(concepto or ''):
            return col
    doc_prefix = doc_code[:2].upper() if doc_code else ''
    if doc_prefix == 'CE':
        return 'CREDITO'
    if doc_prefix == 'CO':
        return 'DEBITO'
    return 'DESCONOCIDO'

def parsear_auxiliar_pdf(ruta, usar_ocr=False):
    texto_completo = ''
    n_pags_ok = 0
    n_pags_mal = 0
    meta = {}

    with pdfplumber.open(ruta) as pdf:
        for pag in pdf.pages:
            t = pag.extract_text() or ''
            if len(t.strip()) <= 30 and usar_ocr and OCR_AVAILABLE:
                t = ocr_pdf_page(ruta, pag.page_number)
            if len(t.strip()) > 30:
                n_pags_ok += 1
                texto_completo += '\n' + t
            else:
                n_pags_mal += 1

    m_si = re.search(r'Saldo\s+Inicial[:\s]+([\d,\.]+)', texto_completo, re.I)
    m_sf = re.search(r'Saldo\s+Final[:\s]+([\d,\.]+)', texto_completo, re.I)
    m_td = re.search(r'Subtotales.*?([\d]{1,3}(?:[,\\.][\d]{3})+(?:\.[\d]+)?)'
                     r'\s+([\d]{1,3}(?:[,\\.][\d]{3})+(?:\.[\d]+)?)',
                     texto_completo, re.I | re.DOTALL)
    meta['SALDO_INICIAL']  = limpiar_num((m_si.group(1) if m_si else '0').replace(',', ''))
    meta['SALDO_FINAL']    = limpiar_num((m_sf.group(1) if m_sf else '0').replace(',', ''))
    if m_td:
        meta['TOTAL_DEBITOS']  = limpiar_num(m_td.group(1).replace(',', ''))
        meta['TOTAL_CREDITOS'] = limpiar_num(m_td.group(2).replace(',', ''))
    else:
        meta['TOTAL_DEBITOS'] = meta['TOTAL_CREDITOS'] = 0
    meta['N_PAGS_OK']  = n_pags_ok
    meta['N_PAGS_MAL'] = n_pags_mal

    # ── Parseo línea a línea (mismo código original) ─────────────────────────
    PAT_DOC    = re.compile(r'^((?:CON|CE|NC)-\d+)\s+(\d{1,2}/\d{1,2}/\d{4})\s+(.*)')
    PAT_MONTO  = re.compile(r'^([\d]{1,3}(?:,[\d]{3})*(?:\.[\d]{1,2})?)$')
    PAT_MPFX   = re.compile(r'^([\d]{1,3}(?:,[\d]{3})*(?:\.[\d]{1,2})?)\s+((?:CON|CE|NC)-\d+.*)$')
    PAT_MSFX   = re.compile(r'\s([\d]{1,3}(?:,[\d]{3})*(?:\.[\d]{1,2})?)$')

    registros   = []
    lineas      = [l.strip() for l in texto_completo.split('\n') if l.strip()]
    pending_doc = None

    def guardar(doc, fecha_str, concepto, monto_str):
        monto = limpiar_num(monto_str.replace(',', ''))
        if not monto or monto <= 0:
            return
        col = determinar_columna(concepto, doc)
        debito  = monto if col == 'DEBITO'  else None
        credito = monto if col == 'CREDITO' else None
        try:
            fdt = pd.to_datetime(fecha_str, format='%d/%m/%Y', errors='coerce')
        except:
            fdt = pd.NaT
        registros.append({
            'DOCUMENTO' : doc,
            'FECHA_RAW' : fecha_str,
            'FECHA'     : fdt,
            'CONCEPTO'  : concepto,
            'DEBITO'    : debito,
            'CREDITO'   : credito,
            'COLUMNA'   : col,
            'VALOR_NETO': (debito or 0) - (credito or 0),
        })

    for linea in lineas:
        m_pfx = PAT_MPFX.match(linea)
        if m_pfx:
            monto_ant = m_pfx.group(1)
            resto     = m_pfx.group(2)
            if pending_doc:
                guardar(pending_doc['doc'], pending_doc['date'],
                        pending_doc['concept'], monto_ant)
                pending_doc = None
            m_doc = PAT_DOC.match(resto)
            if m_doc:
                doc_c   = m_doc.group(1)
                fecha_s = m_doc.group(2)
                concepto_raw = m_doc.group(3)
                m_end = PAT_MSFX.search(concepto_raw)
                if m_end:
                    monto_end = m_end.group(1)
                    concepto_limpio = concepto_raw[:m_end.start()].strip()
                    guardar(doc_c, fecha_s, concepto_limpio, monto_end)
                else:
                    pending_doc = {'doc': doc_c, 'date': fecha_s, 'concept': concepto_raw}
            continue

        m_doc = PAT_DOC.match(linea)
        if m_doc:
            if pending_doc:
                pass
            doc_c   = m_doc.group(1)
            fecha_s = m_doc.group(2)
            concepto_raw = m_doc.group(3)
            m_end = PAT_MSFX.search(concepto_raw)
            if m_end:
                monto_end = m_end.group(1)
                concepto_limpio = concepto_raw[:m_end.start()].strip()
                guardar(doc_c, fecha_s, concepto_limpio, monto_end)
            else:
                pending_doc = {'doc': doc_c, 'date': fecha_s, 'concept': concepto_raw}
            continue

        m_num = PAT_MONTO.match(linea)
        if m_num and pending_doc:
            guardar(pending_doc['doc'], pending_doc['date'],
                    pending_doc['concept'], m_num.group(1))
            pending_doc = None
            continue

        if pending_doc:
            m_end = PAT_MSFX.search(linea)
            if m_end:
                monto_end = m_end.group(1)
                try:
                    mval = float(monto_end.replace(',',''))
                    if mval > 100:
                        guardar(pending_doc['doc'], pending_doc['date'],
                                pending_doc['concept'], monto_end)
                        pending_doc = None
                except:
                    pass

    df = pd.DataFrame(registros)
    if not df.empty:
        df = df.drop_duplicates(subset=['DOCUMENTO','FECHA_RAW','DEBITO','CREDITO'])
        df = df.sort_values('FECHA', na_position='last').reset_index(drop=True)
        df.index += 1
    # Si los totales no se detectaron, calcular desde el dataframe
    if not meta['TOTAL_DEBITOS']:
        meta['TOTAL_DEBITOS'] = df['DEBITO'].sum() if not df.empty else 0
    if not meta['TOTAL_CREDITOS']:
        meta['TOTAL_CREDITOS'] = df['CREDITO'].sum() if not df.empty else 0
    return df, meta

# ── Parseo para CSV/Excel/TXT ─────────────────────────────────────────────
def parsear_banco_csv(df):
    registros = []
    resumen = {}
    # Detectar columnas por nombre
    col_fecha = next((c for c in df.columns if 'fecha' in c.lower()), None)
    col_desc  = next((c for c in df.columns if 'descrip' in c.lower() or 'concepto' in c.lower()), None)
    col_valor = next((c for c in df.columns if 'valor' in c.lower() or 'monto' in c.lower()), None)
    col_saldo = next((c for c in df.columns if 'saldo' in c.lower()), None)

    if not col_fecha: raise ValueError("No se encontró columna de fecha en el archivo del banco")
    for _, row in df.iterrows():
        fecha = str(row[col_fecha])
        try:
            fecha_dt = pd.to_datetime(fecha, dayfirst=True, errors='coerce')
        except:
            fecha_dt = pd.NaT
        desc = str(row[col_desc]) if col_desc else ''
        valor = limpiar_num(row[col_valor]) if col_valor else 0
        saldo = limpiar_num(row[col_saldo]) if col_saldo else None
        registros.append({
            'FECHA_RAW': fecha, 'FECHA': fecha_dt,
            'DESCRIPCION': desc, 'VALOR': valor, 'SALDO': saldo,
            'TIPO': 'ABONO' if (valor or 0) >= 0 else 'CARGO'
        })
    df_out = pd.DataFrame(registros)
    df_out = df_out[df_out['VALOR'].notna()]
    df_out['VALOR'] = pd.to_numeric(df_out['VALOR'], errors='coerce')
    df_out = df_out.drop_duplicates()
    df_out = df_out.sort_values('FECHA', na_position='last').reset_index(drop=True)
    df_out.index += 1
    # Calcular totales
    resumen['TOTAL_ABONOS'] = df_out[df_out['VALOR'] > 0]['VALOR'].sum()
    resumen['TOTAL_CARGOS'] = df_out[df_out['VALOR'] < 0]['VALOR'].sum()
    if col_saldo:
        resumen['SALDO_INICIAL'] = df_out.iloc[0]['SALDO'] if not df_out.empty else 0
        resumen['SALDO_FINAL']   = df_out.iloc[-1]['SALDO'] if not df_out.empty else 0
    else:
        resumen['SALDO_INICIAL'] = resumen['SALDO_FINAL'] = 0
    return df_out, resumen

def _col(df, *palabras):
    """Busca columna por palabras clave (normaliza acentos para comparación)."""
    import unicodedata
    def norm(s):
        return unicodedata.normalize('NFKD', s.lower()).encode('ascii', 'ignore').decode()
    palabras_norm = [norm(p) for p in palabras]
    return next((c for c in df.columns if any(p in norm(c) for p in palabras_norm)), None)

def parsear_auxiliar_csv(df):
    registros = []
    meta = {}
    col_doc = _col(df, 'documento', 'doc')
    col_fec = _col(df, 'fecha')
    col_con = _col(df, 'concepto', 'descrip')
    col_deb = _col(df, 'debito', 'débito', 'debe', 'debitos', 'débitos')
    col_cre = _col(df, 'credito', 'crédito', 'haber', 'creditos', 'créditos')

    if not col_fec: raise ValueError("No se encontró columna de fecha en el auxiliar")

    # Filtrar filas que no sean movimientos reales (sin doc o sin fecha válida)
    for _, row in df.iterrows():
        doc = str(row[col_doc]).strip() if col_doc else ''
        fecha = str(row[col_fec]).strip()
        if not fecha or fecha in ('nan', 'NaT', 'Fecha'): continue
        try: fecha_dt = pd.to_datetime(fecha, dayfirst=True, errors='coerce')
        except: fecha_dt = pd.NaT
        if pd.isna(fecha_dt): continue
        concepto = str(row[col_con]).strip() if col_con else ''
        debito  = limpiar_num(row[col_deb])  if col_deb else None
        credito = limpiar_num(row[col_cre]) if col_cre else None
        # Ignorar filas de subtotales/encabezados de cuenta (sin doc real)
        if not re.match(r'^[A-Z]{2,3}-\d+', doc) and not doc:
            continue
        col_asiento = determinar_columna(concepto, doc)
        registros.append({
            'DOCUMENTO': doc, 'FECHA_RAW': fecha, 'FECHA': fecha_dt,
            'CONCEPTO': concepto, 'DEBITO': debito, 'CREDITO': credito,
            'COLUMNA': col_asiento, 'VALOR_NETO': (debito or 0) - (credito or 0)
        })
    df_out = pd.DataFrame(registros)
    if not df_out.empty:
        # CSV es exportación autoritativa del sistema contable; no deduplicar
        # (misma cuenta puede tener múltiples líneas idénticas en un comprobante)
        df_out = df_out.sort_values('FECHA', na_position='last').reset_index(drop=True)
        df_out.index += 1
    meta['TOTAL_DEBITOS']  = df_out['DEBITO'].sum()  if not df_out.empty else 0
    meta['TOTAL_CREDITOS'] = df_out['CREDITO'].sum() if not df_out.empty else 0
    # Buscar saldo inicial en el CSV (fila con 'Saldo Inicial:')
    if 'SALDO_INICIAL' not in meta:
        meta['SALDO_INICIAL'] = 0
    if 'SALDO_FINAL' not in meta:
        meta['SALDO_FINAL'] = meta['SALDO_INICIAL'] + meta['TOTAL_DEBITOS'] - meta['TOTAL_CREDITOS']
    return df_out, meta

def parsear_banco_txt(texto):
    registros = []
    resumen = {}
    for linea in texto.split('\n'):
        partes = linea.strip().split()
        if not partes or not es_fecha_banco(partes[0]): continue
        fecha_raw = partes[0]
        nums = []; desc_p = []
        for p in partes[1:]:
            v = limpiar_num(p)
            if v is not None: nums.append(v)
            elif not nums: desc_p.append(p)
        if not nums: continue
        saldo = nums[-1]
        valor = nums[-2] if len(nums) >= 2 else nums[0]
        registros.append({
            'FECHA_RAW': fecha_raw,
            'FECHA': pd.to_datetime('2025/' + fecha_raw, format='%Y/%d/%m', errors='coerce'),
            'DESCRIPCION': ' '.join(desc_p), 'VALOR': valor, 'SALDO': saldo,
            'TIPO': 'ABONO' if (valor or 0) >= 0 else 'CARGO'
        })
    df = pd.DataFrame(registros)
    if not df.empty:
        df = df[df['VALOR'].notna()]
        df['VALOR'] = pd.to_numeric(df['VALOR'], errors='coerce')
        df = df.drop_duplicates()
        df = df.sort_values('FECHA', na_position='last').reset_index(drop=True)
        df.index += 1
    resumen['TOTAL_ABONOS'] = df[df['VALOR'] > 0]['VALOR'].sum()
    resumen['TOTAL_CARGOS'] = df[df['VALOR'] < 0]['VALOR'].sum()
    resumen['SALDO_INICIAL'] = df.iloc[0]['SALDO'] if not df.empty else 0
    resumen['SALDO_FINAL']   = df.iloc[-1]['SALDO'] if not df.empty else 0
    return df, resumen

def parsear_auxiliar_txt(texto_completo):
    meta = {}
    registros = []
    lineas = [l.strip() for l in texto_completo.split('\n') if l.strip()]
    pending_doc = None
    PAT_DOC = re.compile(r'^((?:CON|CE|NC)-\d+)\s+(\d{1,2}/\d{1,2}/\d{4})\s+(.*)')
    PAT_MONTO = re.compile(r'^([\d]{1,3}(?:,[\d]{3})*(?:\.[\d]{1,2})?)$')
    PAT_MPFX  = re.compile(r'^([\d]{1,3}(?:,[\d]{3})*(?:\.[\d]{1,2})?)\s+((?:CON|CE|NC)-\d+.*)$')
    PAT_MSFX  = re.compile(r'\s([\d]{1,3}(?:,[\d]{3})*(?:\.[\d]{1,2})?)$')
    def guardar(doc, fecha_s, concepto, monto_str):
        monto = limpiar_num(monto_str.replace(',',''))
        if not monto or monto <= 0: return
        col = determinar_columna(concepto, doc)
        debito = monto if col=='DEBITO' else None
        credito = monto if col=='CREDITO' else None
        try: fdt = pd.to_datetime(fecha_s, format='%d/%m/%Y', errors='coerce')
        except: fdt = pd.NaT
        registros.append({
            'DOCUMENTO': doc, 'FECHA_RAW': fecha_s, 'FECHA': fdt,
            'CONCEPTO': concepto, 'DEBITO': debito, 'CREDITO': credito,
            'COLUMNA': col, 'VALOR_NETO': (debito or 0) - (credito or 0)
        })
    for linea in lineas:
        m_pfx = PAT_MPFX.match(linea)
        if m_pfx:
            if pending_doc:
                guardar(pending_doc['doc'], pending_doc['date'], pending_doc['concept'], m_pfx.group(1))
                pending_doc = None
            m_doc = PAT_DOC.match(m_pfx.group(2))
            if m_doc:
                doc_c, fecha_s, concepto_raw = m_doc.group(1), m_doc.group(2), m_doc.group(3)
                m_end = PAT_MSFX.search(concepto_raw)
                if m_end:
                    guardar(doc_c, fecha_s, concepto_raw[:m_end.start()].strip(), m_end.group(1))
                else:
                    pending_doc = {'doc': doc_c, 'date': fecha_s, 'concept': concepto_raw}
            continue
        m_doc = PAT_DOC.match(linea)
        if m_doc:
            doc_c, fecha_s, concepto_raw = m_doc.group(1), m_doc.group(2), m_doc.group(3)
            m_end = PAT_MSFX.search(concepto_raw)
            if m_end:
                guardar(doc_c, fecha_s, concepto_raw[:m_end.start()].strip(), m_end.group(1))
            else:
                pending_doc = {'doc': doc_c, 'date': fecha_s, 'concept': concepto_raw}
            continue
        if PAT_MONTO.match(linea) and pending_doc:
            guardar(pending_doc['doc'], pending_doc['date'], pending_doc['concept'], linea)
            pending_doc = None
            continue
        if pending_doc:
            m_end = PAT_MSFX.search(linea)
            if m_end:
                try:
                    if float(m_end.group(1).replace(',','')) > 100:
                        guardar(pending_doc['doc'], pending_doc['date'], pending_doc['concept'], m_end.group(1))
                        pending_doc = None
                except: pass
    df = pd.DataFrame(registros)
    df = df.drop_duplicates()
    df = df.sort_values('FECHA', na_position='last').reset_index(drop=True)
    df.index += 1
    meta['TOTAL_DEBITOS'] = df['DEBITO'].sum() if not df.empty else 0
    meta['TOTAL_CREDITOS'] = df['CREDITO'].sum() if not df.empty else 0
    # Saldos no disponibles en TXT generalmente
    meta['SALDO_INICIAL'] = meta['SALDO_FINAL'] = 0
    return df, meta

# ══════════════════════════════════════════════════════════════════════════════
# REGISTRO EXTENSIBLE DE FORMATOS
# ──────────────────────────────────────────────────────────────────────────────
# Cómo añadir un nuevo formato:
#   1. Escribir fn_detectar(ruta, muestra_texto) -> float  [0.0–1.0 confianza]
#   2. Escribir fn_parsear(ruta, usar_ocr)       -> (DataFrame, dict_meta)
#   3. Agregar una entrada al final de REGISTRO_FORMATOS con tipo/ext correctos.
#   No tocar nada más.
# ══════════════════════════════════════════════════════════════════════════════

import unicodedata as _ud

def _norm(s):
    return _ud.normalize('NFKD', s.lower()).encode('ascii', 'ignore').decode()

def _muestra_texto(ruta, ext, n_lineas=50):
    """Texto de muestra para detección rápida (sin parsear el archivo completo)."""
    try:
        if ext == '.pdf':
            with pdfplumber.open(ruta) as pdf:
                return (pdf.pages[0].extract_text() or '') if pdf.pages else ''
        else:
            with open(ruta, 'r', encoding='latin1', errors='replace') as f:
                return ''.join(f.readlines()[:n_lineas])
    except Exception:
        return ''

def _header_row_csv(ruta, encoding='latin1'):
    """Fila donde empieza el encabezado real del CSV (salta metadatos)."""
    claves = {'documento', 'fecha', 'concepto', 'debito', 'credito',
              'valor', 'saldo', 'descripcion'}
    with open(ruta, 'r', encoding=encoding, errors='replace') as f:
        for i, linea in enumerate(f):
            hits = sum(1 for k in claves if k in _norm(linea))
            if hits >= 3:
                return i
    return 0

def _leer_csv_inteligente(ruta):
    """Lee CSV saltando metadatos y capturando saldo inicial si existe."""
    skip = _header_row_csv(ruta)
    saldo_ini = None
    with open(ruta, 'r', encoding='latin1', errors='replace') as f:
        for linea in f.readlines()[:skip + 3]:
            m = re.search(r'Saldo\s+Inicial[:\s]+([\d,\.]+)', linea, re.I)
            if m:
                saldo_ini = limpiar_num(m.group(1).replace(',', ''))
                break
    df = pd.read_csv(ruta, encoding='latin1', sep=None, engine='python',
                     skiprows=skip, header=0)
    df = df.dropna(how='all').reset_index(drop=True)
    return df, saldo_ini

# ── Detectores ────────────────────────────────────────────────────────────────

def _det_bancolombia_pdf(ruta, m):
    hits = [
        bool(re.search(r'ESTADO\s+DE\s+CUENTA',      m, re.I)),
        bool(re.search(r'SALDO\s+ANTERIOR',           m, re.I)),
        bool(re.search(r'TOTAL\s+ABONOS',             m, re.I)),
        bool(re.search(r'TOTAL\s+CARGOS',             m, re.I)),
        bool(re.search(r'BANCOLOMBIA',                m, re.I)),
    ]
    return sum(hits) / len(hits)

def _det_siigo_aux_csv(ruta, m):
    hits = [
        bool(re.search(r'Auxiliares\s*[-–]\s*Plan\s+de\s+Cuentas', m, re.I)),
        bool(re.search(r'Nit\.?\s*\d{6,}',            m, re.I)),
        bool(re.search(r'(?:CE|CON|NC)-\d+',          m)),
        bool(re.search(r'D[eé]bitos?',                m, re.I)),
        bool(re.search(r'Cr[eé]ditos?',               m, re.I)),
        bool(re.search(r'Saldo\s+Inicial',            m, re.I)),
    ]
    return sum(hits) / len(hits)

def _det_helisa_aux_csv(ruta, m):
    hits = [
        bool(re.search(r'HELISA',                     m, re.I)),
        bool(re.search(r'Libro\s+Auxiliar',           m, re.I)),
        bool(re.search(r'(?:CE|CON|NC|RE|RG)-\d+',   m)),
        bool(re.search(r'D[eé]bito|Cr[eé]dito',      m, re.I)),
    ]
    return sum(hits) / len(hits)

def _det_world_office_aux_csv(ruta, m):
    hits = [
        bool(re.search(r'World\s*Office|WO\s+\d',    m, re.I)),
        bool(re.search(r'Comprobante',                m, re.I)),
        bool(re.search(r'D[eé]bito|Cr[eé]dito',      m, re.I)),
        bool(re.search(r'\d{4}-\d{2}-\d{2}',         m)),
    ]
    return sum(hits) / len(hits)

def _det_aux_pdf_generico(ruta, m):
    hits = [
        bool(re.search(r'(?:CON|CE|NC|RE|RG)-\d+',   m)),
        bool(re.search(r'Saldo\s+Inicial',            m, re.I)),
        bool(re.search(r'Subtotales',                 m, re.I)),
    ]
    return sum(hits) / len(hits)

def _det_banco_csv_generico(ruta, m):
    hits = [
        bool(re.search(r'\bfecha\b',                  m, re.I)),
        bool(re.search(r'\bvalor\b|\bmonto\b',        m, re.I)),
        bool(re.search(r'\bsaldo\b',                  m, re.I)),
    ]
    return sum(hits) / len(hits)

def _det_aux_csv_generico(ruta, m):
    hits = [
        bool(re.search(r'\bdocumento\b',              m, re.I)),
        bool(re.search(r'\bfecha\b',                  m, re.I)),
        bool(re.search(r'\bconcepto\b|\bdescripci',   m, re.I)),
        bool(re.search(r'\bd[eé]bito\b|\bhaber\b',   m, re.I)),
    ]
    return sum(hits) / len(hits)

def _det_banco_txt(ruta, m):
    lineas_fecha = sum(1 for l in m.split('\n')
                       if re.match(r'^\d{1,2}/\d{2}\s', l.strip()))
    return min(1.0, lineas_fecha / 5)

def _det_aux_txt(ruta, m):
    hits = [
        bool(re.search(r'(?:CON|CE|NC)-\d+',         m)),
        bool(re.search(r'\d{1,2}/\d{2}/\d{4}',       m)),
    ]
    return sum(hits) / len(hits)

# ── Parsers de formato ─────────────────────────────────────────────────────────

def _par_bancolombia_pdf(ruta, usar_ocr):
    return parsear_banco_pdf(ruta, usar_ocr=usar_ocr)

def _par_siigo_aux_csv(ruta, usar_ocr):
    df_raw, saldo_ini = _leer_csv_inteligente(ruta)
    df, meta = parsear_auxiliar_csv(df_raw)
    if saldo_ini is not None:
        meta['SALDO_INICIAL'] = saldo_ini
        meta['SALDO_FINAL'] = saldo_ini + meta.get('TOTAL_DEBITOS',0) - meta.get('TOTAL_CREDITOS',0)
    return df, meta

def _par_helisa_aux_csv(ruta, usar_ocr):
    # Helisa exporta similar a SIIGO; reutilizamos misma lógica
    return _par_siigo_aux_csv(ruta, usar_ocr)

def _par_world_office_aux_csv(ruta, usar_ocr):
    # World Office: columnas pueden llamarse "Debe"/"Haber" en lugar de Débito/Crédito
    df_raw, saldo_ini = _leer_csv_inteligente(ruta)
    df, meta = parsear_auxiliar_csv(df_raw)
    if saldo_ini is not None:
        meta['SALDO_INICIAL'] = saldo_ini
        meta['SALDO_FINAL'] = saldo_ini + meta.get('TOTAL_DEBITOS',0) - meta.get('TOTAL_CREDITOS',0)
    return df, meta

def _par_aux_pdf_generico(ruta, usar_ocr):
    return parsear_auxiliar_pdf(ruta, usar_ocr=usar_ocr)

def _par_banco_csv_generico(ruta, usar_ocr):
    df_raw, _ = _leer_csv_inteligente(ruta)
    return parsear_banco_csv(df_raw)

def _par_aux_csv_generico(ruta, usar_ocr):
    df_raw, saldo_ini = _leer_csv_inteligente(ruta)
    df, meta = parsear_auxiliar_csv(df_raw)
    if saldo_ini is not None:
        meta['SALDO_INICIAL'] = saldo_ini
        meta['SALDO_FINAL'] = saldo_ini + meta.get('TOTAL_DEBITOS',0) - meta.get('TOTAL_CREDITOS',0)
    return df, meta

def _par_banco_txt(ruta, usar_ocr):
    with open(ruta, 'r', encoding='utf-8', errors='replace') as f:
        return parsear_banco_txt(f.read())

def _par_aux_txt(ruta, usar_ocr):
    with open(ruta, 'r', encoding='utf-8', errors='replace') as f:
        return parsear_auxiliar_txt(f.read())

# ── Registro principal ────────────────────────────────────────────────────────
# Orden importa: formatos más específicos primero dentro del mismo tipo/ext.
# El despachador elige el de mayor confianza, no el primero.

REGISTRO_FORMATOS = [
    # ── BANCO ─────────────────────────────────────────────────────────────────
    {
        'nombre'  : 'Bancolombia — PDF Estado de Cuenta',
        'tipo'    : 'BANCO',
        'ext'     : ['.pdf'],
        'detectar': _det_bancolombia_pdf,
        'parsear' : _par_bancolombia_pdf,
    },
    {
        'nombre'  : 'Banco — CSV/Excel con encabezados estándar',
        'tipo'    : 'BANCO',
        'ext'     : ['.csv', '.xlsx', '.xls'],
        'detectar': _det_banco_csv_generico,
        'parsear' : _par_banco_csv_generico,
    },
    {
        'nombre'  : 'Banco — TXT líneas fecha/descripción/valor',
        'tipo'    : 'BANCO',
        'ext'     : ['.txt'],
        'detectar': _det_banco_txt,
        'parsear' : _par_banco_txt,
    },
    # ── AUXILIAR ──────────────────────────────────────────────────────────────
    {
        'nombre'  : 'SIIGO — CSV Auxiliares Plan de Cuentas',
        'tipo'    : 'AUXILIAR',
        'ext'     : ['.csv'],
        'detectar': _det_siigo_aux_csv,
        'parsear' : _par_siigo_aux_csv,
    },
    {
        'nombre'  : 'Helisa — CSV Libro Auxiliar',
        'tipo'    : 'AUXILIAR',
        'ext'     : ['.csv', '.xlsx', '.xls'],
        'detectar': _det_helisa_aux_csv,
        'parsear' : _par_helisa_aux_csv,
    },
    {
        'nombre'  : 'World Office — CSV Auxiliar',
        'tipo'    : 'AUXILIAR',
        'ext'     : ['.csv', '.xlsx', '.xls'],
        'detectar': _det_world_office_aux_csv,
        'parsear' : _par_world_office_aux_csv,
    },
    {
        'nombre'  : 'Auxiliar Contable — PDF (CON/CE/NC)',
        'tipo'    : 'AUXILIAR',
        'ext'     : ['.pdf'],
        'detectar': _det_aux_pdf_generico,
        'parsear' : _par_aux_pdf_generico,
    },
    {
        'nombre'  : 'Auxiliar Contable — CSV/Excel genérico',
        'tipo'    : 'AUXILIAR',
        'ext'     : ['.csv', '.xlsx', '.xls'],
        'detectar': _det_aux_csv_generico,
        'parsear' : _par_aux_csv_generico,
    },
    {
        'nombre'  : 'Auxiliar Contable — TXT',
        'tipo'    : 'AUXILIAR',
        'ext'     : ['.txt'],
        'detectar': _det_aux_txt,
        'parsear' : _par_aux_txt,
    },
]

# ── Despachador ───────────────────────────────────────────────────────────────

def _despachar(ruta, tipo, ext, usar_ocr):
    """Elige el mejor formato registrado y lo parsea. Retorna (df, meta, nombre_formato, confianza)."""
    candidatos = [f for f in REGISTRO_FORMATOS if f['tipo'] == tipo and ext in f['ext']]
    if not candidatos:
        raise ValueError(f"Formato no soportado para tipo={tipo}, extensión={ext}")
    muestra = _muestra_texto(ruta, ext)
    puntuaciones = [(f, f['detectar'](ruta, muestra)) for f in candidatos]
    mejor, conf = max(puntuaciones, key=lambda x: x[1])
    df, meta = mejor['parsear'](ruta, usar_ocr)
    return df, meta, mejor['nombre'], round(conf * 100)

# ── Función unificada de carga ────────────────────────────────────────────────
def cargar_y_parsear(uploaded_file, tipo, usar_ocr=False):
    nombre = uploaded_file.name
    ext = Path(nombre).suffix.lower()
    with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
        tmp.write(uploaded_file.getvalue())
        ruta = tmp.name

    try:
        if ext == '.pdf':
            diag = diagnosticar_pdf(ruta, tipo)
            ocr_efectivo = usar_ocr or diag['ocr_usado']
            df, res, fmt_nombre, fmt_conf = _despachar(ruta, tipo, ext, ocr_efectivo)
            legibilidad = (diag['pct_estimado_datos'], diag['calidad'],
                           diag['advertencias'], fmt_nombre, fmt_conf)
        elif ext in ['.csv', '.xlsx', '.xls']:
            df, res, fmt_nombre, fmt_conf = _despachar(ruta, tipo, ext, usar_ocr)
            legibilidad = (100.0, '🟢 EXCELENTE', [], fmt_nombre, fmt_conf)
        elif ext == '.txt':
            df, res, fmt_nombre, fmt_conf = _despachar(ruta, tipo, ext, usar_ocr)
            calidad = '🟢 EXCELENTE' if not df.empty else '🟠 PARCIAL'
            pct     = 95.0 if not df.empty else 50.0
            adv     = [] if not df.empty else ['Archivo TXT sin datos reconocibles']
            legibilidad = (pct, calidad, adv, fmt_nombre, fmt_conf)
        else:
            raise ValueError(f"Formato no soportado: {ext}")
    except Exception as e:
        raise e
    finally:
        os.unlink(ruta)
    return df, res, legibilidad

# ── Comparación (original) ─────────────────────────────────────────────────
TOL_EXACTA = 1.0
TOL_APROX  = 0.005

def comparar_documentos(df_b, df_a):
    if df_b.empty or df_a.empty:
        return pd.DataFrame(), df_a.copy() if not df_a.empty else pd.DataFrame()
    idx_usados = set()
    filas = []
    for idx_b, row_b in df_b.iterrows():
        vb = row_b['VALOR']
        if pd.isna(vb): continue
        monto_abs = abs(vb)
        col_buscar = 'DEBITO' if vb >= 0 else 'CREDITO'
        libres = df_a[df_a.index.map(lambda i: i not in idx_usados) & df_a[col_buscar].notna()].copy()
        match_tipo = match_monto = match_idx = None
        match_doc = match_conc = match_fecha_aux = ''
        if not libres.empty:
            libres['_diff'] = (libres[col_buscar] - monto_abs).abs()
            exactos = libres[libres['_diff'] <= TOL_EXACTA]
            if not exactos.empty:
                mejor = exactos.nsmallest(1, '_diff').iloc[0]
                match_tipo = 'EXACTO'; match_monto = mejor[col_buscar]; match_idx = mejor.name
                match_doc = mejor['DOCUMENTO']; match_conc = mejor['CONCEPTO']; match_fecha_aux = mejor['FECHA_RAW']
            if match_tipo is None and monto_abs > 0:
                aprox = libres[libres['_diff'] / monto_abs <= TOL_APROX]
                if not aprox.empty:
                    mejor = aprox.nsmallest(1, '_diff').iloc[0]
                    match_tipo = 'APROX'; match_monto = mejor[col_buscar]; match_idx = mejor.name
                    match_doc = mejor['DOCUMENTO']; match_conc = mejor['CONCEPTO']; match_fecha_aux = mejor['FECHA_RAW']
        if match_idx is not None: idx_usados.add(match_idx)
        estado = '✅ COINCIDE EXACTO' if match_tipo=='EXACTO' else '🔶 COINCIDE APROX.' if match_tipo=='APROX' else '❌ SOLO EN BANCO'
        diff_val = abs(monto_abs - match_monto) if match_monto is not None else None
        filas.append({
            'N': idx_b, 'FECHA_BANCO': row_b['FECHA_RAW'], 'TIPO_MOV': row_b['TIPO'],
            'DESCRIPCION': row_b['DESCRIPCION'], 'VALOR_BANCO': vb,
            'DOC_AUXILIAR': match_doc, 'FECHA_AUXILIAR': match_fecha_aux,
            'CONCEPTO_AUX': match_conc, 'MONTO_AUXILIAR': match_monto,
            'DIFERENCIA': diff_val, 'ESTADO': estado, 'MATCH_TIPO': match_tipo or 'SIN_MATCH'
        })
    df_comp = pd.DataFrame(filas)
    df_solo_aux = df_a[~df_a.index.isin(idx_usados)].copy()
    df_solo_aux['ESTADO'] = '📋 SOLO EN AUXILIAR'
    return df_comp, df_solo_aux

# ── Interfaz Streamlit ──────────────────────────────────────────────────────
st.title("🏦 CREDIEXPRESS POPAYÁN SAS — Conciliación Bancaria Premium")
st.markdown("### Extracto Bancolombia ↔ Auxiliar Contable · Cuenta 1120.05.01")
st.markdown("**Enero 2025** · Cuenta Ahorros #26100001167")

with st.sidebar:
    st.header("📂 Cargar Archivos")
    st.info("**Formatos aceptados:** PDF (incluye OCR automático si es escaneado), CSV, Excel (.xlsx), TXT")
    banco_file = st.file_uploader("Extracto Bancolombia", type=["pdf","csv","xlsx","txt"])
    aux_file   = st.file_uploader("Auxiliar Contable", type=["pdf","csv","xlsx","txt"])
    usar_ocr = st.checkbox("Forzar OCR en PDF escaneados (si está instalado)", value=True)
    ejecutar = st.button("🚀 Ejecutar análisis completo", disabled=not (banco_file and aux_file))
    if ejecutar:
        st.session_state.run = True

    st.markdown("---")
    st.markdown("### 💡 Recomendaciones para máxima precisión")
    st.markdown("""
    - **Extracto bancario:** PDF original del banco (no escaneado) o CSV con columnas `Fecha, Descripción, Valor, Saldo`.
    - **Auxiliar contable:** Preferiblemente **CSV o Excel** exportado desde SIIGO/Helisa/World Office con columnas `Documento, Fecha, Concepto, Debito, Credito`.
    - Si el PDF está escaneado (imagen), instale **Tesseract OCR** y **Poppler** para extracción automática.
    """)
    if OCR_AVAILABLE:
        st.success("✅ OCR está disponible (se usará automáticamente en páginas sin texto)")
    else:
        st.warning("⚠️ OCR no instalado. Los PDF escaneados no se podrán leer automáticamente.\n"
                   "Instale 'pytesseract', 'pdf2image' y el motor Tesseract+Poppler.")

if 'run' in st.session_state and st.session_state.run:
    with st.spinner("Procesando archivos..."):
        try:
            df_banco, res_banco, leg_banco = cargar_y_parsear(banco_file, 'BANCO', usar_ocr=usar_ocr)
            sa  = res_banco.get('SALDO_INICIAL', 0) or res_banco.get('SALDO_ANTERIOR', 0) or 0
            sac = res_banco.get('SALDO_FINAL', 0) or res_banco.get('SALDO_ACTUAL', 0) or 0
            tab_s = res_banco.get('TOTAL_ABONOS', 0) or 0
            tca_s = abs(res_banco.get('TOTAL_CARGOS', 0)) or 0  # cargos negativos en CSV los ponemos absolutos

            df_aux, meta_aux, leg_aux = cargar_y_parsear(aux_file, 'AUXILIAR', usar_ocr=usar_ocr)
            si_a = meta_aux.get('SALDO_INICIAL', 0) or 0
            sf_a = meta_aux.get('SALDO_FINAL',   0) or 0
            td_a = meta_aux.get('TOTAL_DEBITOS', 0) or 0
            tc_a = meta_aux.get('TOTAL_CREDITOS',0) or 0
        except Exception as e:
            st.error(f"❌ Error al procesar los archivos: {e}")
            st.stop()

    # Comparación
    if not df_aux.empty:
        df_comp, df_solo_aux = comparar_documentos(df_banco, df_aux)
        n_tot  = len(df_comp)
        n_exac = (df_comp['ESTADO'] == '✅ COINCIDE EXACTO').sum()
        n_apr  = (df_comp['ESTADO'] == '🔶 COINCIDE APROX.').sum()
        n_sbco = (df_comp['ESTADO'] == '❌ SOLO EN BANCO').sum()
        n_saux = len(df_solo_aux)
        pct_conc = (n_exac + n_apr) / max(n_tot, 1) * 100
        exactas = df_comp[df_comp['ESTADO'] == '✅ COINCIDE EXACTO']
        aprox   = df_comp[df_comp['ESTADO'] == '🔶 COINCIDE APROX.']
        s_banco = df_comp[df_comp['ESTADO'] == '❌ SOLO EN BANCO']
    else:
        df_comp = pd.DataFrame()
        df_solo_aux = pd.DataFrame()
        n_tot = n_exac = n_apr = n_sbco = n_saux = pct_conc = 0
        exactas = aprox = s_banco = pd.DataFrame()

    # ── Pestañas ──────────────────────────────────────────────────────────
    tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs([
        "📊 Diagnóstico", "🏦 Extracto Banco", "📋 Auxiliar Contable",
        "🔗 Comparación", "📝 Diferencias", "⚖️ Conciliación Formal",
        "📈 Visualizaciones", "💾 Exportar Excel"
    ])

    with tab1:
        st.header("Diagnóstico de Legibilidad")
        # leg_banco / leg_aux = (pct, calidad, advertencias, fmt_nombre, fmt_confianza)
        p_banco, cal_banco, adv_banco, fmt_banco, conf_banco = leg_banco
        p_aux,   cal_aux,   adv_aux,   fmt_aux,   conf_aux   = leg_aux

        c1, c2 = st.columns(2)
        with c1:
            st.subheader("Extracto Banco")
            st.metric("Legibilidad", f"{p_banco:.1f}%", cal_banco)
            st.info(f"**Formato detectado:** {fmt_banco}  \n**Confianza:** {conf_banco}%")
            for a in adv_banco: st.warning(a)
        with c2:
            st.subheader("Auxiliar Contable")
            st.metric("Legibilidad", f"{p_aux:.1f}%", cal_aux)
            st.info(f"**Formato detectado:** {fmt_aux}  \n**Confianza:** {conf_aux}%")
            for a in adv_aux: st.warning(a)
        if p_banco >= 95 and p_aux >= 95:
            st.success("✅ Ambos archivos completamente legibles.")
        elif p_banco < 80 or p_aux < 80:
            st.error("⚠️ Legibilidad baja. Si es PDF escaneado, active el OCR o convierta a CSV/Excel.")

    with tab2:
        st.header("Extracto Bancolombia")
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Saldo Inicial", cop(sa))
        col2.metric("Total Abonos (+)", cop(tab_s))
        col3.metric("Total Cargos (-)", cop(tca_s))
        col4.metric("Saldo Final", cop(sac))
        dif_arit = (sa + tab_s - tca_s) - sac
        st.write(f"Verificación aritmética: {cop(sa+tab_s-tca_s)} calculado")
        st.write(f"Saldo declarado: {cop(sac)}")
        if abs(dif_arit) < 1:
            st.success(f"Diferencia: {cop(dif_arit)} ✅ CUADRA")
        else:
            st.error(f"Diferencia: {cop(dif_arit)} ⚠️ REVISAR")
        st.subheader("Primeras 25 transacciones")
        st.dataframe(df_banco[['FECHA_RAW','DESCRIPCION','VALOR','SALDO','TIPO']].head(25),
                     use_container_width=True)

    with tab3:
        st.header("Auxiliar Contable")
        if not df_aux.empty:
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Saldo Inicial", cop(si_a))
            col2.metric("Total Débitos", cop(td_a))
            col3.metric("Total Créditos", cop(tc_a))
            col4.metric("Saldo Final", cop(sf_a))
            dif_arit_aux = (si_a + td_a - tc_a) - sf_a
            st.write(f"Verificación: {cop(si_a+td_a-tc_a)} calculado vs {cop(sf_a)} declarado")
            if abs(dif_arit_aux) < 1:
                st.success(f"Diferencia: {cop(dif_arit_aux)} ✅ CUADRA")
            else:
                st.error(f"Diferencia: {cop(dif_arit_aux)} ⚠️ REVISAR")
            deb_df = df_aux[df_aux['DEBITO'].notna()]
            cre_df = df_aux[df_aux['CREDITO'].notna()]
            des_df = df_aux[df_aux['COLUMNA'] == 'DESCONOCIDO']
            st.write(f"Asientos DÉBITO: {len(deb_df)} — {cop(deb_df['DEBITO'].sum())}")
            st.write(f"Asientos CRÉDITO: {len(cre_df)} — {cop(cre_df['CREDITO'].sum())}")
            if len(des_df) > 0:
                st.warning(f"Asientos sin clasificar: {len(des_df)}")
            st.subheader("Primeras 25 líneas")
            st.dataframe(df_aux[['DOCUMENTO','FECHA_RAW','CONCEPTO','DEBITO','CREDITO','COLUMNA']].head(25),
                         use_container_width=True)
        else:
            st.error("No se extrajeron asientos del auxiliar.")

    with tab4:
        st.header("Comparación Uno a Uno")
        if df_aux.empty:
            st.warning("Sin datos")
        else:
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Analizados", n_tot)
            c2.metric("✅ Exactos", n_exac, f"{n_exac/max(n_tot,1)*100:.0f}%")
            c3.metric("🔶 Aprox.", n_apr, f"{n_apr/max(n_tot,1)*100:.0f}%")
            c4.metric("❌ Solo Banco", n_sbco, f"{n_sbco/max(n_tot,1)*100:.0f}%")
            st.progress(pct_conc/100, text=f"Tasa de conciliación: {pct_conc:.1f}%")
            st.metric("📋 Solo Auxiliar", n_saux)
            st.subheader("Tabla Completa (movimiento bancario vs auxiliar)")
            st.dataframe(df_comp, use_container_width=True)

    with tab5:
        st.header("Reporte Detallado de Diferencias")
        if df_aux.empty:
            st.warning("Sin datos")
        else:
            with st.expander("✅ Coincidencias Exactas", expanded=True):
                st.write(f"**{len(exactas)} movimientos conciliados**")
                st.write(f"Total valor banco: {cop(exactas['VALOR_BANCO'].sum())}")
                st.dataframe(exactas[['FECHA_BANCO','TIPO_MOV','VALOR_BANCO','DOC_AUXILIAR','MONTO_AUXILIAR']].head(60),
                             use_container_width=True)
                if len(exactas) > 60: st.caption(f"... y {len(exactas)-60} más (ver Excel)")

            with st.expander("🔶 Coincidencias Aproximadas"):
                if aprox.empty:
                    st.write("(ninguna)")
                else:
                    st.dataframe(aprox[['FECHA_BANCO','TIPO_MOV','VALOR_BANCO','MONTO_AUXILIAR','DIFERENCIA','DOC_AUXILIAR']],
                                 use_container_width=True)

            with st.expander("❌ Movimientos Bancarios sin Registro Contable"):
                if s_banco.empty:
                    st.success("Todos los movimientos tienen asiento contable.")
                else:
                    st.write(f"**Valor total sin registro:** {cop(s_banco['VALOR_BANCO'].sum())}")
                    abonos_sb = s_banco[s_banco['VALOR_BANCO'] > 0]
                    cargos_sb = s_banco[s_banco['VALOR_BANCO'] < 0]
                    st.write(f"Abonos sin asiento: {cop(abonos_sb['VALOR_BANCO'].sum())} ({len(abonos_sb)} trans.)")
                    st.write(f"Cargos sin asiento: {cop(cargos_sb['VALOR_BANCO'].sum())} ({len(cargos_sb)} trans.)")
                    st.dataframe(s_banco[['FECHA_BANCO','TIPO_MOV','VALOR_BANCO','DESCRIPCION']],
                                 use_container_width=True)
                    st.caption("Causas probables: movimientos de fin de mes, intereses, GMF, comisiones, pendientes contables.")

            with st.expander("📋 Asientos Auxiliar sin Transacción Bancaria"):
                if df_solo_aux.empty:
                    st.success("Todos los asientos tienen transacción bancaria.")
                else:
                    st.write(f"Débitos sin banco: {cop(df_solo_aux['DEBITO'].sum())}")
                    st.write(f"Créditos sin banco: {cop(df_solo_aux['CREDITO'].sum())}")
                    st.dataframe(df_solo_aux[['FECHA_RAW','DOCUMENTO','DEBITO','CREDITO','CONCEPTO']],
                                 use_container_width=True)
                    st.caption("Causas probables: asientos de cierre, pagos en efectivo, notas de ajuste internas.")

    with tab6:
        st.header("Conciliación Bancaria Formal (Formato Estándar Colombia)")
        st.markdown("---")
        st.subheader("I. Saldo según Extracto Bancario")
        st.text(f"Saldo anterior (31/12/2024)        {cop(sa)}")
        st.text(f"(+) Total abonos                      {cop(tab_s)}")
        st.text(f"(-) Total cargos                      {cop(tca_s)}")
        calc_banco = sa + tab_s - tca_s
        dif_b = calc_banco - sac
        st.text(f"{'─'*60}")
        st.text(f"(=) Saldo calculado                   {cop(calc_banco)}")
        st.text(f"(=) Saldo declarado (31/01/2025)      {cop(sac)}")
        if abs(dif_b) < 1: st.success(f"Diferencia: {cop(dif_b)} ✅ CUADRA")
        else: st.error(f"Diferencia: {cop(dif_b)} ⚠️ REVISAR")

        st.markdown("---")
        st.subheader("II. Saldo según Auxiliar Contable (Cuenta 1120.05.01)")
        st.text(f"Saldo inicial (01/01/2025)        {cop(si_a)}")
        st.text(f"(+) Total Débitos                     {cop(td_a)}")
        st.text(f"(-) Total Créditos                    {cop(tc_a)}")
        calc_aux = si_a + td_a - tc_a
        dif_a = calc_aux - sf_a
        st.text(f"{'─'*60}")
        st.text(f"(=) Saldo calculado                   {cop(calc_aux)}")
        st.text(f"(=) Saldo final declarado (31/01)     {cop(sf_a)}")
        if abs(dif_a) < 1: st.success(f"Diferencia: {cop(dif_a)} ✅ CUADRA")
        else: st.error(f"Diferencia: {cop(dif_a)} ⚠️ REVISAR")

        st.markdown("---")
        st.subheader("III. Diferencias Banco ↔ Auxiliar")
        dif_saldos = sac - sf_a
        st.text(f"Saldo banco (31/01):                  {cop(sac)}")
        st.text(f"Saldo auxiliar (31/01):               {cop(sf_a)}")
        st.text(f"{'─'*60}")
        st.text(f"DIFERENCIA NETA DE SALDOS:            {cop(dif_saldos)}")
        st.text(f"Abonos banco vs Débitos auxiliar:     {cop(tab_s - td_a)}")
        st.text(f"Cargos banco vs Créditos auxiliar:    {cop(tca_s - tc_a)}")

        if not df_aux.empty:
            st.markdown("---")
            st.subheader("IV. Composición de la Diferencia (Análisis de Conciliación)")
            val_sin_aux  = s_banco['VALOR_BANCO'].sum() if not s_banco.empty else 0
            val_sin_banco_deb = df_solo_aux['DEBITO'].sum()  if not df_solo_aux.empty else 0
            val_sin_banco_cre = df_solo_aux['CREDITO'].sum() if not df_solo_aux.empty else 0
            st.text(f"Monto banco sin registro auxiliar:    {cop(val_sin_aux)}")
            st.text(f"Monto aux. Débito sin banco:          {cop(val_sin_banco_deb)}")
            st.text(f"Monto aux. Crédito sin banco:         {cop(val_sin_banco_cre)}")
            st.text(f"Tasa de conciliación:                 {pct_conc:.1f}%")

        st.markdown("---")
        st.markdown("**CARLOS ANDRÉS SILVA VELA**    **FERNANDO CUCALÓN SÁNCHEZ**")
        st.markdown("REPRESENTANTE LEGAL         CONTADOR")
        st.markdown("C.C. 1061717925             T.P. 23049-T")

    with tab7:
        st.header("Visualizaciones")
        plt.rcParams.update({'font.family': 'DejaVu Sans', 'figure.dpi': 110})
        fig, axes = plt.subplots(2, 3, figsize=(22, 12))
        fig.suptitle('CREDIEXPRESS POPAYAN SAS — Conciliación Bancaria Enero 2025',
                     fontsize=14, fontweight='bold', y=1.01)

        # G1
        ax1 = axes[0, 0]
        df_s = df_banco[df_banco['SALDO'].notna()].copy()
        if not df_s.empty:
            ax1.plot(range(len(df_s)), df_s['SALDO']/1e6, color='#1565C0', lw=1.2)
            ax1.fill_between(range(len(df_s)), df_s['SALDO']/1e6, alpha=0.12, color='#1565C0')
        ax1.set_title('Evolución del Saldo Bancario', fontweight='bold')
        ax1.set_ylabel('Millones COP')
        ax1.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'${x:.0f}M'))
        ax1.grid(True, alpha=0.3)

        # G2
        ax2 = axes[0, 1]
        if n_tot > 0:
            cont = df_comp['ESTADO'].value_counts()
            color_map = {
                '✅ COINCIDE EXACTO': '#4CAF50',
                '🔶 COINCIDE APROX.': '#FFC107',
                '❌ SOLO EN BANCO':   '#F44336',
            }
            cs = [color_map.get(e, '#9E9E9E') for e in cont.index]
            lbl = [e + ' (' + str(v) + ')' for e, v in zip(cont.index, cont.values)]
            ax2.pie(cont.values, labels=lbl, colors=cs, autopct='%1.0f%%', startangle=90,
                    textprops={'fontsize': 8})
        ax2.set_title('Estado Conciliacion', fontweight='bold')

        # G3
        ax3 = axes[0, 2]
        cats  = ['Entradas\nBanco', 'Debitos\nAuxiliar', 'Salidas\nBanco', 'Creditos\nAuxiliar']
        vals3 = [tab_s/1e6, td_a/1e6, tca_s/1e6, tc_a/1e6]
        cols3 = ['#2196F3', '#4CAF50', '#F44336', '#FF9800']
        bars3 = ax3.bar(cats, vals3, color=cols3, alpha=0.85)
        ax3.set_title('Totales: Banco vs Auxiliar', fontweight='bold')
        ax3.set_ylabel('Millones COP')
        ax3.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'${x:.0f}M'))
        for b, v in zip(bars3, vals3):
            ax3.text(b.get_x() + b.get_width()/2, b.get_height() + 1,
                     f'${v:.0f}M', ha='center', fontsize=8, fontweight='bold')
        ax3.grid(True, axis='y', alpha=0.3)

        # G4
        ax4 = axes[1, 0]
        if not df_banco.empty:
            df_banco['DIA'] = df_banco['FECHA_RAW'].apply(
                lambda x: int(str(x).split('/')[0]) if '/' in str(x) else 0)
            por_dia = df_banco.groupby(['DIA', 'TIPO'])['VALOR'].sum().unstack(fill_value=0)
            if 'ABONO' in por_dia.columns:
                ax4.bar(por_dia.index, por_dia['ABONO']/1e6,
                        label='Abonos (+)', color='#2E7D32', alpha=0.8)
            if 'CARGO' in por_dia.columns:
                ax4.bar(por_dia.index, por_dia['CARGO'].abs()/1e6,
                        label='Cargos (-)', color='#C62828', alpha=0.7)
        ax4.set_title('Movimientos por Dia', fontweight='bold')
        ax4.set_xlabel('Dia de Enero')
        ax4.set_ylabel('Millones COP')
        ax4.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'${x:.0f}M'))
        ax4.legend(fontsize=8)
        ax4.grid(True, axis='y', alpha=0.3)

        # G5
        ax5 = axes[1, 1]
        if not df_aux.empty:
            tipo_cnt = df_aux['DOCUMENTO'].str[:2].value_counts()
            cols5 = ['#4CAF50', '#2196F3', '#FF9800'][:len(tipo_cnt)]
            bars5 = ax5.bar(tipo_cnt.index, tipo_cnt.values, color=cols5)
            ax5.set_title('Asientos por Tipo (Auxiliar)', fontweight='bold')
            ax5.set_ylabel('N asientos')
            for b, v in zip(bars5, tipo_cnt.values):
                ax5.text(b.get_x() + b.get_width()/2, b.get_height() + 1,
                         str(v), ha='center', fontsize=9, fontweight='bold')
            ax5.grid(True, axis='y', alpha=0.3)
        else:
            ax5.text(0.5, 0.5, 'Sin datos auxiliar', ha='center', va='center', fontsize=12)
            ax5.set_title('Asientos por Tipo (Auxiliar)', fontweight='bold')

        # G6
        ax6 = axes[1, 2]
        if n_tot > 0:
            ve = exactas['VALOR_BANCO'].abs().sum()/1e6 if not exactas.empty else 0
            va = aprox['VALOR_BANCO'].abs().sum()/1e6 if not aprox.empty else 0
            vs = s_banco['VALOR_BANCO'].abs().sum()/1e6 if not s_banco.empty else 0
            vx = (df_solo_aux['DEBITO'].fillna(0).sum() + df_solo_aux['CREDITO'].fillna(0).sum())/1e6 if not df_solo_aux.empty else 0
            lbl6 = ['Exacto', 'Aprox.', 'Solo\nbanco', 'Solo\nauxiliar']
            val6 = [ve, va, vs, vx]
            col6 = ['#4CAF50', '#FFC107', '#F44336', '#2196F3']
            bars6 = ax6.bar(lbl6, val6, color=col6, alpha=0.85)
            ax6.set_title('Valor por Estado Conciliacion', fontweight='bold')
            ax6.set_ylabel('Millones COP')
            ax6.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'${x:.0f}M'))
            for b, v in zip(bars6, val6):
                ax6.text(b.get_x() + b.get_width()/2, b.get_height() + 0.5,
                         f'${v:.1f}M', ha='center', fontsize=8, fontweight='bold')
            ax6.grid(True, axis='y', alpha=0.3)

        plt.tight_layout()
        st.pyplot(fig)

    with tab8:
        st.header("Exportar a Excel (8 Hojas Premium)")
        nombre_salida = 'CREDIEXPRESS_Conciliacion_Enero2025.xlsx'

        FILL_VERDE    = PatternFill('solid', fgColor='C8F7C5')
        FILL_AMARILLO = PatternFill('solid', fgColor='FFF3CD')
        FILL_ROJO     = PatternFill('solid', fgColor='F7C5C5')
        FILL_AZUL     = PatternFill('solid', fgColor='D0E8FF')
        FILL_HEADER   = PatternFill('solid', fgColor='1565C0')
        FONT_HEADER   = Font(bold=True, color='FFFFFF', size=10)

        def estilizar_hoja(ws):
            for cell in ws[1]:
                cell.fill = FILL_HEADER
                cell.font = FONT_HEADER
                cell.alignment = Alignment(horizontal='center', vertical='center')
            for col in ws.columns:
                mx = max((len(str(c.value or '')) for c in col), default=10)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 3, 55)

        def colorear_por_estado(ws, col_estado_idx):
            for row in ws.iter_rows(min_row=2):
                val = str(row[col_estado_idx - 1].value or '')
                fill = (FILL_VERDE    if '✅' in val else
                        FILL_AMARILLO if '🔶' in val else
                        FILL_ROJO     if '❌' in val else
                        FILL_AZUL     if '📋' in val else None)
                if fill:
                    for cell in row:
                        cell.fill = fill

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Hoja 1
            if not df_aux.empty:
                h1 = df_comp[['N','FECHA_BANCO','TIPO_MOV','DESCRIPCION','VALOR_BANCO',
                               'DOC_AUXILIAR','FECHA_AUXILIAR','CONCEPTO_AUX','MONTO_AUXILIAR',
                               'DIFERENCIA','ESTADO']].copy()
                h1.columns = ['N','Fecha_Banco','Tipo','Descripcion_Banco','Valor_Banco',
                              'Doc_Auxiliar','Fecha_Auxiliar','Concepto_Auxiliar','Monto_Auxiliar',
                              'Diferencia','Estado']
            else:
                h1 = pd.DataFrame({'Info': ['Sin comparación']})
            h1.to_excel(writer, sheet_name='1_Comparacion_Completa', index=False)

            for estado, nombre in [
                ('✅ COINCIDE EXACTO', '2_Coincidencias_Exactas'),
                ('🔶 COINCIDE APROX.', '3_Coincidencias_Aprox'),
                ('❌ SOLO EN BANCO',   '4_Solo_Banco_Sin_Auxiliar'),
            ]:
                sub = df_comp[df_comp['ESTADO'] == estado].copy() if not df_aux.empty else pd.DataFrame()
                if sub.empty: sub = pd.DataFrame({'Info': ['Sin registros']})
                sub.to_excel(writer, sheet_name=nombre, index=False)

            if not df_solo_aux.empty:
                df_solo_aux.to_excel(writer, sheet_name='5_Solo_Auxiliar_Sin_Banco', index=False)
            else:
                pd.DataFrame({'Info': ['Todos los asientos tienen movimiento bancario']}).to_excel(
                    writer, sheet_name='5_Solo_Auxiliar_Sin_Banco', index=False)

            df_banco.to_excel(writer, sheet_name='6_Extracto_Banco_Completo', index=True)
            df_aux.to_excel(writer, sheet_name='7_Auxiliar_Contable_Completo', index=True)

            resumen_data = {
                'Concepto': [
                    'Archivo banco', 'Archivo auxiliar',
                    'Saldo inicial banco', 'Saldo final banco',
                    'Total abonos banco', 'Total cargos banco',
                    'Saldo inicial auxiliar', 'Saldo final auxiliar',
                    'Total debitos auxiliar', 'Total creditos auxiliar',
                    'Diferencia saldos finales',
                    'Movimientos analizados', 'Coincidencias exactas',
                    'Coincidencias aprox.', 'Solo en banco', 'Solo en auxiliar',
                    'Tasa de conciliacion %',
                ],
                'Valor': [
                    banco_file.name, aux_file.name,
                    sa, sac, tab_s, tca_s,
                    si_a, sf_a, td_a, tc_a,
                    sac - sf_a,
                    n_tot, n_exac, n_apr, n_sbco, n_saux,
                    round(pct_conc, 1),
                ]
            }
            pd.DataFrame(resumen_data).to_excel(writer, sheet_name='8_Resumen_Conciliacion', index=False)

            wb = writer.book
            for sname in wb.sheetnames:
                ws = wb[sname]
                estilizar_hoja(ws)
            if '1_Comparacion_Completa' in wb.sheetnames and not df_aux.empty:
                ws1 = wb['1_Comparacion_Completa']
                colorear_por_estado(ws1, 11)

        output.seek(0)
        st.download_button(
            label="📥 Descargar Excel Premium",
            data=output,
            file_name=nombre_salida,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    st.success("✅ Conciliación completada exitosamente")